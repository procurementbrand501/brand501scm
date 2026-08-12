import os
import warnings
import pandas as pd
import numpy as np
import json
import glob
import re
import openpyxl
from datetime import datetime, timedelta

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def find_latest_inventory_file(prefix):
    """폴더에서 prefix로 시작하는 .xlsx 파일 중 날짜(YYYYMMDD)가 가장 큰 것 반환"""
    files = glob.glob(f"{prefix}*.xlsx")
    if not files:
        return None
    def extract_date(fname):
        m = re.search(r'(\d{8})', fname)
        return int(m.group(1)) if m else 0
    return max(files, key=extract_date)


def find_latest_sales_file(keywords):
    """국가별 판매 파일(*SCM_재고관리*.xlsx) 중 국가 키워드가 포함된 파일을 모아
    수정시각(최종 저장 시간)이 가장 최신인 것을 반환.
    같은 이름으로 재다운로드하면 '(1)' 복사본이 생기므로 파일명이 아니라 mtime 기준."""
    candidates = [
        f for f in glob.glob("*SCM_재고관리*.xlsx")
        if not os.path.basename(f).startswith("~$")
        and any(kw in f for kw in keywords)
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def load_robust_excel(file_path, target_keyword='품목코드', sheet_name=0):
    if not os.path.exists(file_path):
        return None
    try:
        df_raw = pd.read_excel(file_path, header=None, sheet_name=sheet_name)
        header_idx = None
        for idx, row in df_raw.iterrows():
            if row.astype(str).str.contains(target_keyword).any():
                header_idx = idx
                break
        if header_idx is not None:
            raw_cols = [str(x).strip() for x in df_raw.iloc[header_idx]]
            columns = []
            seen = {}
            for c in raw_cols:
                if c in seen:
                    seen[c] += 1
                    columns.append(f"{c}_{seen[c]}")
                else:
                    seen[c] = 0
                    columns.append(c)
            df_refined = df_raw.iloc[header_idx + 1:].copy()
            df_refined.columns = columns
            return df_refined.reset_index(drop=True)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df.columns = [str(c).strip() for c in df.columns]
            return df
    except Exception as e:
        print(f"❌ '{file_path}' 시트 '{sheet_name}' 로드 중 에러: {e}")
        return None


def parse_int(val):
    if pd.notna(val) and str(val).strip() != '-':
        try:
            return int(float(str(val).strip().replace(',', '')))
        except:
            return 0
    return 0


def find_sku_col(df):
    """재고 파일들의 SKU 컬럼명 탐색. 2026-07-10 FBA 파일부터 'Seller SKU'가 'SKU'로 변경됨
    (FNSKU 컬럼이 추가되며 명칭 정리된 것으로 보임) — 일반/FBT/FBA 세 파일 모두 같은 변경
    리스크가 있어 공통 헬퍼로 둠."""
    for c in ('Seller SKU', 'SKU'):
        if c in df.columns:
            return c
    raise KeyError("'Seller SKU' 또는 'SKU' 컬럼을 찾을 수 없습니다.")


def assign_region(warehouse_name):
    wh_str = str(warehouse_name)
    if '미국' in wh_str:
        return 'us'
    elif '영국' in wh_str:
        return 'uk'
    elif any(keyword in wh_str for keyword in ['독일', '네덜란드', '유럽', 'EU']):
        return 'eu'
    return '기타'


def read_first_matching_sheet(path, required_cols):
    """여러 시트 중 required_cols를 모두 포함한 첫 시트를 DataFrame으로 반환.
    재고 파일에 피벗 시트가 첫 번째로 저장된 경우 대응 (2026-07-09 일반_전체재고 파일에서 발생 —
    첫 시트가 '행 레이블' 피벗이라 기존 pd.read_excel(첫 시트 고정)이 조용히 실패)."""
    xls = pd.ExcelFile(path)
    for sn in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sn)
            df.columns = [str(c).strip() for c in df.columns]
            if all(c in df.columns for c in required_cols):
                return df
        except Exception:
            continue
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_integrated_inventory(file_gen, file_fbt, file_fba):
    """일반/FBT/FBA 통합 재고 파일 3종 읽어 국가별 창고재고 합계 산출"""
    all_rows = []

    def expand_multicode_rows(df_input):
        expanded = []
        for _, row in df_input.iterrows():
            code_str = str(row['품목코드']).strip()
            if ',' in code_str:
                sub_codes = [c.strip() for c in code_str.split(',') if c.strip()]
                for sc in sub_codes:
                    new_row = row.copy()
                    new_row['품목코드'] = sc
                    expanded.append(new_row)
            else:
                expanded.append(row)
        return pd.DataFrame(expanded).reset_index(drop=True)

    if file_gen and os.path.exists(file_gen):
        try:
            df_gen = read_first_matching_sheet(file_gen, ['창고', '가용 재고 (개)'])
            tmp = pd.DataFrame({
                '창고': df_gen['창고'],
                '품목코드': df_gen[find_sku_col(df_gen)].astype(str).str.strip().str.upper(),
                '가용재고': df_gen['가용 재고 (개)'].apply(parse_int)
            })
            tmp = expand_multicode_rows(tmp)
            all_rows.append(tmp)
            print(f"   > {file_gen} 로드 완료 ({len(tmp)}행)")
        except Exception as e:
            print(f"   > ⚠️ {file_gen} 처리 중 에러: {e}")

    if file_fbt and os.path.exists(file_fbt):
        try:
            df_fbt = read_first_matching_sheet(file_fbt, ['가용 재고 (개)'])
            tmp = pd.DataFrame({
                '창고': df_fbt['창고'],
                '품목코드': df_fbt[find_sku_col(df_fbt)].astype(str).str.strip().str.upper(),
                '가용재고': df_fbt['가용 재고 (개)'].apply(parse_int)
            })
            tmp = expand_multicode_rows(tmp)
            all_rows.append(tmp)
            print(f"   > {file_fbt} 로드 완료 ({len(tmp)}행)")
        except Exception as e:
            print(f"   > ⚠️ {file_fbt} 처리 중 에러: {e}")

    if file_fba and os.path.exists(file_fba):
        try:
            df_fba = read_first_matching_sheet(file_fba, ['가용 재고 (개)'])
            warehouse_col = '창고' if '창고' in df_fba.columns else (
                '창고재고그룹' if '창고재고그룹' in df_fba.columns else None
            )
            if warehouse_col is None:
                raise KeyError("'창고' 또는 '창고재고그룹' 컬럼을 찾을 수 없습니다.")
            tmp = pd.DataFrame({
                '창고': df_fba[warehouse_col],
                '품목코드': df_fba[find_sku_col(df_fba)].astype(str).str.strip().str.upper(),
                '가용재고': df_fba['가용 재고 (개)'].apply(parse_int)
            })
            tmp = expand_multicode_rows(tmp)
            all_rows.append(tmp)
            print(f"   > {file_fba} 로드 완료 ({len(tmp)}행)")
        except Exception as e:
            print(f"   > ⚠️ {file_fba} 처리 중 에러: {e}")

    if not all_rows:
        return pd.DataFrame(columns=['품목코드', 'us_stock', 'uk_stock', 'eu_stock'])

    df_total = pd.concat(all_rows, ignore_index=True)
    df_total = df_total.dropna(subset=['품목코드'])
    df_total['품목코드'] = df_total['품목코드'].astype(str).str.strip()
    df_total['region'] = df_total['창고'].apply(assign_region)
    df_total = df_total[df_total['region'].isin(['us', 'uk', 'eu'])]

    df_grouped = df_total.groupby(['품목코드', 'region'])['가용재고'].sum().reset_index()
    df_pivot = df_grouped.pivot(index='품목코드', columns='region', values='가용재고').fillna(0).astype(int).reset_index()

    for region in ['us', 'uk', 'eu']:
        if region not in df_pivot.columns:
            df_pivot[region] = 0

    df_pivot = df_pivot.rename(columns={'us': 'us_stock', 'uk': 'uk_stock', 'eu': 'eu_stock'})
    return df_pivot[['품목코드', 'us_stock', 'uk_stock', 'eu_stock']]


def parse_kr_available(file_gen):
    """일반_전체재고 파일에서 국내(콜로세움) 창고의 SKU별 가용재고(F열) 추출.
    부킹 관리에서 '한국에서 보낼 수 있는 재고'의 베이스로 사용 (권역 현지 재고와 별개).
    """
    if not file_gen or not os.path.exists(file_gen):
        return {}
    try:
        df = read_first_matching_sheet(file_gen, ['창고', '가용 재고 (개)'])
        sku_col = find_sku_col(df)
    except Exception as e:
        print(f"   > ⚠️ 한국(콜로세움) 가용재고 파싱 실패: {e}")
        return {}
    if '창고' not in df.columns or '가용 재고 (개)' not in df.columns:
        print("   > ⚠️ 한국(콜로세움) 가용재고: 필요한 컬럼(창고/가용 재고)을 찾지 못함")
        return {}

    kr_map = {}
    for _, row in df.iterrows():
        if '콜로세움' not in str(row['창고']):
            continue
        code_str = str(row[sku_col]).strip().upper()
        if not code_str or code_str == 'NAN':
            continue
        qty = parse_int(row['가용 재고 (개)'])
        # 멀티코드(콤마 구분)는 기존 재고 로드와 동일하게 각 코드에 동일 수량 부여
        codes = [c.strip() for c in code_str.split(',')] if ',' in code_str else [code_str]
        for c in codes:
            if c:
                kr_map[c] = kr_map.get(c, 0) + qty
    return kr_map


def read_and_summarize_sales(file_path, region_prefix):
    if not os.path.exists(file_path):
        print(f"⚠️ '{file_path}' 파일이 없어 {region_prefix} 일판매 0 처리")
        return pd.DataFrame(columns=['품목코드', f'{region_prefix}_sales'])
    try:
        df_raw = pd.read_excel(file_path, header=None)
    except Exception as e:
        print(f"❌ '{file_path}' 로드 에러: {e}")
        return pd.DataFrame(columns=['품목코드', f'{region_prefix}_sales'])

    sales_col_idx = None
    for r_idx in range(min(20, len(df_raw) - 1)):
        row_values = df_raw.iloc[r_idx].tolist()
        next_row_values = df_raw.iloc[r_idx + 1].tolist()
        for c_idx in range(min(len(row_values), len(next_row_values))):
            val_current = row_values[c_idx]
            val_next = next_row_values[c_idx]
            if pd.notna(val_current) and '당월 일평균 판매량' in str(val_current):
                if pd.notna(val_next) and '소계' in str(val_next):
                    sales_col_idx = c_idx
                    break
        if sales_col_idx is not None:
            break

    df = load_robust_excel(file_path, '품목코드')
    if df is None or '품목코드' not in df.columns:
        return pd.DataFrame(columns=['품목코드', f'{region_prefix}_sales'])

    df = df.dropna(subset=['품목코드'])
    df['품목코드'] = df['품목코드'].astype(str).str.strip()
    df = df[df['품목코드'] != '품목코드']

    explicit_sales_col = next((c for c in df.columns if '일평균' in str(c) and '판매' in str(c)), None)
    if explicit_sales_col:
        df[f'{region_prefix}_sales'] = df[explicit_sales_col].apply(parse_int)
    elif sales_col_idx is not None and sales_col_idx < len(df.columns):
        try:
            df[f'{region_prefix}_sales'] = df.iloc[:, sales_col_idx].apply(parse_int)
        except Exception:
            df[f'{region_prefix}_sales'] = 0
    else:
        df[f'{region_prefix}_sales'] = 0

    return df[['품목코드', f'{region_prefix}_sales']]


def _find_daily_layout(df_raw):
    """판매 시트에서 일자별 컬럼(datetime 헤더 셀)과 품목코드 컬럼 위치를 탐색.
    파일마다 헤더 행 위치가 다르므로(US 3행째, UK/EU 2행째) 앞쪽 행들을 스캔해서 찾는다.
    날짜 datetime 셀이 있는 그 컬럼이 해당 일자 블록의 '소계' 컬럼이다.
    """
    date_cols = []
    for r in range(min(6, len(df_raw))):
        row = df_raw.iloc[r].tolist()
        cols = [(i, v) for i, v in enumerate(row) if isinstance(v, datetime)]
        if len(cols) >= 5:  # 날짜 컬럼이 여러 개 있는 행이 일자별 헤더 (단일 datetime 셀 오탐 방지)
            date_cols = [(i, v.date()) for i, v in cols]
            break
    code_row, code_col = None, None
    for r in range(min(8, len(df_raw))):
        for i, v in enumerate(df_raw.iloc[r].tolist()):
            if str(v).strip() == '품목코드':
                code_row, code_col = r, i
                break
        if code_col is not None:
            break
    return date_cols, code_row, code_col


def parse_daily_sales(file_path):
    """판매 파일의 최신 2개 월 시트('26년 7월' 형식)에서 SKU별 일자별 판매량(소계) 추출.
    반환: {품목코드: {date: qty}}. 월 경계를 넘는 28일 롤링 평균 계산을 위해 두 달치를 합친다.
    """
    if not os.path.exists(file_path):
        return {}
    try:
        xls = pd.ExcelFile(file_path)
    except Exception:
        return {}

    monthly_sheets = []
    for s in xls.sheet_names:
        m = re.match(r'(\d{2})년\s*(\d{1,2})월$', str(s).strip())
        if m:
            monthly_sheets.append((2000 + int(m.group(1)), int(m.group(2)), s))
    monthly_sheets.sort(reverse=True)

    result = {}
    for year, month, sheet in monthly_sheets[:2]:
        try:
            df_raw = pd.read_excel(file_path, sheet_name=sheet, header=None)
        except Exception:
            continue
        date_cols, code_row, code_col = _find_daily_layout(df_raw)
        if not date_cols or code_col is None:
            print(f"   > ⚠️ '{sheet}' 일자별 컬럼 탐색 실패 — 이 시트는 건너뜀")
            continue
        for r in range(code_row + 1, len(df_raw)):
            code = df_raw.iat[r, code_col]
            if pd.isna(code):
                continue
            code = str(code).strip().upper()
            if not code or code == '품목코드':
                continue
            entry = result.setdefault(code, {})
            for ci, d in date_cols:
                if ci >= df_raw.shape[1]:
                    continue
                qty = parse_int(df_raw.iat[r, ci])
                if qty:
                    entry[d] = entry.get(d, 0) + qty
    return result


def summarize_sales_28d(daily_map, region_prefix, today=None):
    """일자별 판매에서 '어제까지' 기준 다중 윈도우(7/14/28일) 평균 중 최댓값을 일판매로 산출.
    - 28일 단일 평균은 판매가 막 시작된 SKU(결품 후 재개, 신규 수요)를 과소평가해서
      항공 긴급을 늦게 감지하는 문제가 있었음 (2026-07-15). max 방식이라 보수적(재고 여유 방향).
    - 어떤 윈도우가 채택됐는지 {r}_sales_window(7/14/28)로 함께 반환 — 셀 툴팁 표시용.
    급증 감지: 최근 3일 평균이 max(10, 28일 평균×2) 이상이면 spike 플래그.
    반환 DataFrame: 품목코드 / {r}_sales / {r}_sales_window / {r}_last7(list) / {r}_spike(bool)
    """
    today = today or datetime.now().date()
    start = today - timedelta(days=28)
    rows = []
    for code, dates in daily_map.items():
        window = [dates.get(start + timedelta(days=i), 0) for i in range(28)]
        avg28 = sum(window) / 28
        avg14 = sum(window[-14:]) / 14
        avg7 = sum(window[-7:]) / 7
        # 동률이면 긴 윈도우 우선 (수치가 안정적이고, 툴팁에 '급증 반영'으로 오해되지 않도록)
        best_avg, best_window = avg28, 28
        if avg14 > best_avg:
            best_avg, best_window = avg14, 14
        if avg7 > best_avg:
            best_avg, best_window = avg7, 7
        recent3 = sum(window[-3:]) / 3
        spike = recent3 >= max(10, 2 * avg28)
        rows.append({
            '품목코드': code,
            f'{region_prefix}_sales': int(round(best_avg)),
            f'{region_prefix}_sales_window': best_window,
            f'{region_prefix}_last7': window[-7:],
            f'{region_prefix}_spike': bool(spike),
        })
    return pd.DataFrame(rows)


def build_group_sales(merged, region_daily, group_col='통합 제품명'):
    """통합 제품명 그룹(tab2)의 일판매를 '멤버 코드의 일자별 판매를 먼저 합산한 뒤' 윈도우 평균으로 산출.

    코드별 {r}_sales를 단순 합산하면 안 되는 이유: 2026-07-15부터 코드마다 채택 윈도우가
    달라질 수 있어(max(7·14·28일)), 구코드의 28일 평균 + 신코드의 7일 평균처럼 서로 다른
    기간의 수요가 더해져 과대 계상된다. 코드 전환 그룹에서 실제 발생
    (GLMX090→GLMX178: 3,094 + 2,151 = 5,245로 표시됐지만 실제 그룹 수요는 3,632).
    일자별로 먼저 합산하면 전환기에도 그룹의 실제 수요 곡선 하나만 남는다.

    반환: {region: DataFrame(group_col, {r}_sales, {r}_sales_window, {r}_last7, {r}_spike)}
          일자별 파싱이 실패한 권역(폴백 경로)은 결과에서 빠지며, 호출부는 기존 합산값을 유지한다.
    """
    out = {}
    if group_col not in merged.columns:
        return out
    # (코드, 그룹) 쌍 단위 중복 제거 — 마스터의 유령 중복 코드('GLMX183' vs 'GLMX183\n')로
    # merge 시 같은 코드 행이 여러 벌 생기면 한 그룹 안에서 같은 SKU 판매가 중복 합산된다.
    # 코드만으로 dedup하면 반대로 다른 그룹에 붙은 같은 코드가 통째로 누락되므로 쌍으로 dedup한다.
    pairs = merged[['품목코드', group_col]].dropna(subset=['품목코드']).copy()
    pairs['_code_key'] = pairs['품목코드'].astype(str).str.strip().str.upper()
    pairs = pairs.drop_duplicates(subset=['_code_key', group_col])
    for r, dmap in region_daily.items():
        if not dmap:
            continue
        group_daily = {}
        for code, gname in zip(pairs['_code_key'], pairs[group_col]):
            d = dmap.get(code)
            if not d:
                continue
            entry = group_daily.setdefault(str(gname), {})
            for dt, q in d.items():
                entry[dt] = entry.get(dt, 0) + q
        if group_daily:
            out[r] = summarize_sales_28d(group_daily, r).rename(columns={'품목코드': group_col})
    return out


def build_group_fcst(master_df, fcst_map):
    """마스터의 코드→통합 제품명 매핑으로 월별 판매예정을 그룹 단위 합산.

    merged 기반 집계(agg_fcst)를 쓰면 안 되는 이유: merged는 재고·판매·이동중 데이터가 있는
    코드로만 만들어지고 마스터는 거기에 left join되므로, **재고가 아직 없는 신코드는 행 자체가
    없어** 그 코드의 포캐스팅 물량이 통째로 누락된다. 제품 리뉴얼로 코드가 바뀌면
    (GLMX124 → GLMX124_V02) 신코드에 잡힌 향후 계획이 사라져 그룹 포캐스팅이 0으로 떨어짐
    (2026-07-22 발견). 마스터를 순회하면 재고 유무와 무관하게 전부 집계된다.

    반환: {통합 제품명: {"YYYY-MM": {"us","uk","eu"}}}
    """
    out = {}
    if master_df is None or '품목코드' not in master_df.columns or '통합 제품명' not in master_df.columns:
        return out
    for _, r in master_df.dropna(subset=['품목코드']).iterrows():
        gname = str(r.get('통합 제품명', '')).strip()
        if not gname or gname.lower() == 'nan':
            continue
        for part in str(r['품목코드']).split(','):
            code = part.strip().upper()
            months = fcst_map.get(code) if code else None
            if not months:
                continue
            g = out.setdefault(gname, {})
            for mk, regs in months.items():
                o = g.setdefault(mk, {'us': 0, 'uk': 0, 'eu': 0})
                for rg in ('us', 'uk', 'eu'):
                    o[rg] += regs.get(rg, 0)
    return out


def read_sales_rolling(file_path, region_prefix):
    """일자별 파싱 기반 28일 롤링 평균으로 일판매 산출. 파싱 실패 시 기존 '당월 일평균' 방식으로 폴백.
    반환: (DataFrame, daily_map) — daily_map은 모달의 일별 추이 차트용.
    """
    daily_map = parse_daily_sales(file_path)
    if not daily_map:
        print(f"   > ⚠️ {region_prefix.upper()}: 일자별 파싱 실패 → 기존 '당월 일평균' 컬럼으로 폴백")
        return read_and_summarize_sales(file_path, region_prefix), {}

    df_new = summarize_sales_28d(daily_map, region_prefix)

    # 검증 로그: 기존 당월 일평균 방식과 비교 (둘 다 판매가 있는 SKU 기준)
    df_old = read_and_summarize_sales(file_path, region_prefix).rename(
        columns={f'{region_prefix}_sales': '_old_sales'})
    cmp = pd.merge(df_new[['품목코드', f'{region_prefix}_sales']], df_old, on='품목코드', how='inner')
    both = cmp[(cmp[f'{region_prefix}_sales'] > 0) & (cmp['_old_sales'] > 0)]
    if len(both) > 0:
        ratio = (both[f'{region_prefix}_sales'] / both['_old_sales']).mean()
        print(f"   > {region_prefix.upper()}: 28일 롤링 {len(df_new)}개 SKU 산출 "
              f"(기존 당월 일평균 대비 평균 비율 {ratio:.2f}, 비교 가능 {len(both)}개)")
    else:
        print(f"   > {region_prefix.upper()}: 28일 롤링 {len(df_new)}개 SKU 산출 (기존 방식과 비교 가능한 SKU 없음)")

    return df_new, daily_map


def read_transit_data(file_path):
    df = load_robust_excel(file_path, '품목코드')
    if df is None or '품목코드' not in df.columns:
        print(f"⚠️ 이동중 재고 파일('{file_path}') 없음")
        return pd.DataFrame(columns=['품목코드', 'us_air', 'us_sea', 'uk_air', 'uk_sea', 'eu_air', 'eu_sea'])
    df = df.dropna(subset=['품목코드'])
    df['품목코드'] = df['품목코드'].astype(str).str.strip()
    mapping = {
        'us_air': 'US항공', 'us_sea': 'US해상',
        'uk_air': 'UK항공', 'uk_sea': 'UK해상',
        'eu_air': 'EU항공', 'eu_sea': 'EU해상'
    }
    res_df = pd.DataFrame({'품목코드': df['품목코드']})
    for key, col_name in mapping.items():
        if col_name in df.columns:
            res_df[key] = df[col_name].apply(parse_int)
        else:
            alt_col = next((c for c in df.columns if key.split('_')[0].upper() in str(c) and ('항공' if 'air' in key else '해상') in str(c)), None)
            res_df[key] = df[alt_col].apply(parse_int) if alt_col else 0
    return res_df


def parse_transit_schedule(file_path, year=2026):
    """ETA 파일에서 SKU별 입고 일정 [{region, mode, date, qty}] 추출"""
    if not os.path.exists(file_path):
        print(f"   > ⚠️ ETA 파일 없음: {file_path}")
        return {}

    schedule_map = {}

    target_sheets = [
        ('us', 'air', 'US항공'),
        ('us', 'sea', 'US해상'),
        ('uk', 'air', 'UK항공'),
        ('uk', 'sea', 'UK해상'),
        ('eu', 'air', 'EU항공'),
        ('eu', 'sea', 'EU해상'),
    ]

    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        print(f"   > ⚠️ ETA 파일 열기 실패: {e}")
        return {}

    for region, mode, sheet_keyword in target_sheets:
        matched_sheet = next((s for s in xls.sheet_names if s.startswith(sheet_keyword)), None)
        if matched_sheet is None:
            print(f"   > ⚠️ '{sheet_keyword}' 시트를 찾을 수 없습니다.")
            continue

        try:
            df_raw = pd.read_excel(file_path, sheet_name=matched_sheet, header=None)
        except Exception as e:
            print(f"   > ⚠️ '{matched_sheet}' 읽기 실패: {e}")
            continue

        header_row = None
        for r in range(min(10, len(df_raw))):
            row_vals = df_raw.iloc[r].tolist()
            if any('품목코드' in str(v) for v in row_vals):
                header_row = r
                break
        if header_row is None:
            print(f"   > ⚠️ '{matched_sheet}'에서 헤더를 찾지 못함")
            continue

        header = df_raw.iloc[header_row].tolist()
        code_col_idx = None
        date_cols = []
        tbd_cols = []   # ETA 미정 차수 컬럼 (예: '해상4차(TBD)', '레오스 특송(TBD)') — 날짜 없이 수량만 있음

        date_pattern = re.compile(r'\((\d{1,2})/(\d{1,2})\)')

        for i, h in enumerate(header):
            if h is None:
                continue
            h_str = str(h)
            if '품목코드' in h_str and code_col_idx is None:
                code_col_idx = i
            m = date_pattern.search(h_str)
            if m:
                month = int(m.group(1))
                day = int(m.group(2))
                try:
                    eta_date = datetime(year, month, day).date()
                    date_cols.append((i, eta_date, h_str.strip()))
                except ValueError:
                    pass
            elif 'TBD' in h_str.upper():
                tbd_cols.append((i, h_str.strip()))

        if code_col_idx is None:
            print(f"   > ⚠️ '{matched_sheet}': 품목코드 컬럼 못 찾음")
            continue

        if date_cols or tbd_cols:
            for r in range(header_row + 1, len(df_raw)):
                code = df_raw.iat[r, code_col_idx]
                if pd.isna(code):
                    continue
                sku = str(code).strip().upper()
                if not sku or sku == 'NAN' or sku == '품목코드':
                    continue

                for col_idx, eta_date, label in date_cols:
                    if col_idx >= df_raw.shape[1]:
                        continue
                    val = df_raw.iat[r, col_idx]
                    qty = parse_int(val)
                    if qty <= 0:
                        continue
                    schedule_map.setdefault(sku, []).append({
                        'region': region,
                        'mode': mode,
                        'date': eta_date.isoformat(),
                        'qty': qty,
                        'label': label
                    })

                for col_idx, label in tbd_cols:
                    if col_idx >= df_raw.shape[1]:
                        continue
                    qty = parse_int(df_raw.iat[r, col_idx])
                    if qty <= 0:
                        continue
                    schedule_map.setdefault(sku, []).append({
                        'region': region,
                        'mode': mode,
                        'date': None,
                        'qty': qty,
                        'label': label
                    })
            print(f"   > '{matched_sheet}' 파싱 완료 (도착일 컬럼 {len(date_cols)}개 + ETA미정 컬럼 {len(tbd_cols)}개)")
            continue

        # 대체 스키마: 여러 날짜 컬럼 대신 '입고 예정일' + '수량' 단일 컬럼 (예: EU항공)
        eta_date_col_idx = next((i for i, h in enumerate(header) if h is not None and '입고 예정일' in str(h)), None)
        qty_col_idx = next((i for i, h in enumerate(header) if h is not None and str(h).strip() == '수량'), None)

        if eta_date_col_idx is None or qty_col_idx is None:
            print(f"   > ⚠️ '{matched_sheet}': 도착일 컬럼(다중 날짜 또는 입고 예정일+수량) 못 찾음")
            continue

        row_count = 0
        for r in range(header_row + 1, len(df_raw)):
            code = df_raw.iat[r, code_col_idx]
            if pd.isna(code):
                continue
            sku = str(code).strip().upper()
            if not sku or sku == 'NAN' or sku == '품목코드':
                continue

            raw_date = df_raw.iat[r, eta_date_col_idx] if eta_date_col_idx < df_raw.shape[1] else None
            if pd.isna(raw_date):
                continue
            try:
                eta_date = pd.to_datetime(raw_date).date()
            except Exception:
                continue

            qty = parse_int(df_raw.iat[r, qty_col_idx]) if qty_col_idx < df_raw.shape[1] else 0
            if qty <= 0:
                continue

            schedule_map.setdefault(sku, []).append({
                'region': region,
                'mode': mode,
                'date': eta_date.isoformat(),
                'qty': qty
            })
            row_count += 1

        print(f"   > '{matched_sheet}' 파싱 완료 (입고 예정일+수량 방식, {row_count}건)")

    print(f"   > ETA 일정 추출 완료: 총 {len(schedule_map)}개 SKU의 입고 기록")
    return schedule_map


def parse_domestic_inbound(file_path):
    """
    국내 메인창고 입고 스케줄 파싱
    파일: 2026_SCM_영업채널별_판매_포캐스팅 / 탭: '입고 스케줄_상세'
    반환: { 품목코드: { 'total': 미입총계, 'items': [{month, week, qty}, ...] } }
    """
    if not os.path.exists(file_path):
        print(f"   > ⚠️ 입고 스케줄 파일 없음: {file_path}")
        return {}

    try:
        df_raw = pd.read_excel(file_path, sheet_name='입고 스케줄_상세', header=None)
    except Exception as e:
        print(f"   > ⚠️ 입고 스케줄 시트 읽기 실패: {e}")
        return {}

    if len(df_raw) < 5:
        print("   > ⚠️ 입고 스케줄 시트 데이터 부족")
        return {}

    row_month = df_raw.iloc[2].tolist()
    row_week  = df_raw.iloc[3].tolist()

    code_col_idx = None
    total_col_idx = None
    week_cols = []  # [(col_idx, '6월', '22~28...')]

    current_month = None
    for i in range(len(row_week)):
        m = row_month[i] if i < len(row_month) else None
        w = row_week[i] if i < len(row_week) else None

        if m is not None and m != '':
            m_str = str(m).strip()
            if m_str == '미입 총계':
                total_col_idx = i
                continue
            if '월' in m_str:
                current_month = m_str

        if w == '품목코드':
            code_col_idx = i
            continue

        if w is not None and isinstance(w, str):
            w_clean = w.replace('\n', ' ').strip()
            if w_clean == '소계':
                continue
            if current_month and ('~' in w_clean or '영업일' in w_clean):
                week_cols.append((i, current_month, w_clean))

    if code_col_idx is None or total_col_idx is None or not week_cols:
        print(f"   > ⚠️ 입고 스케줄 헤더 파싱 실패 (code={code_col_idx}, total={total_col_idx}, weeks={len(week_cols)})")
        return {}

    result = {}
    for r in range(4, len(df_raw)):
        row = df_raw.iloc[r].tolist()
        if code_col_idx >= len(row):
            continue
        code = row[code_col_idx]
        if pd.isna(code):
            continue
        code = str(code).strip().upper()
        if not code or code in ('NAN', '품목코드'):
            continue

        total_val = parse_int(row[total_col_idx]) if total_col_idx < len(row) else 0
        items = []
        for col_idx, month, week in week_cols:
            if col_idx >= len(row):
                continue
            qty = parse_int(row[col_idx])
            if qty > 0:
                items.append({'month': month, 'week': week, 'qty': qty})

        if total_val > 0 or items:
            result[code] = {'total': total_val, 'items': items}

    print(f"   > 입고 스케줄 파싱 완료: {len(result)}개 SKU")
    return result


def parse_product_cost(file_path, sheet_name='입고 스케줄_상세'):
    """입고 스케줄_상세 시트에서 품목코드별 원가(E열 '원가(vat+)') 추출 → 재고 금액 계산에 사용"""
    if not os.path.exists(file_path):
        return {}
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        print(f"   > ⚠️ 원가 파싱 실패: {e}")
        return {}

    header_row = None
    for r in range(min(6, len(df_raw))):
        if any('품목코드' in str(v) for v in df_raw.iloc[r].tolist()):
            header_row = r
            break
    if header_row is None:
        print("   > ⚠️ 원가 시트에서 헤더(품목코드)를 찾지 못함")
        return {}

    header = df_raw.iloc[header_row].tolist()
    code_col = next((i for i, h in enumerate(header) if str(h).strip() == '품목코드'), None)
    cost_col = next((i for i, h in enumerate(header) if '원가' in str(h)), None)
    if code_col is None or cost_col is None:
        print(f"   > ⚠️ 원가 컬럼 못 찾음 (code={code_col}, cost={cost_col})")
        return {}

    cost_map = {}
    for r in range(header_row + 1, len(df_raw)):
        row = df_raw.iloc[r].tolist()
        if code_col >= len(row):
            continue
        code = row[code_col]
        if pd.isna(code):
            continue
        code = str(code).strip().upper()
        if not code or code == '품목코드':
            continue
        cost = parse_int(row[cost_col]) if cost_col < len(row) else 0
        if cost > 0 and code not in cost_map:
            cost_map[code] = cost
    return cost_map


def parse_product_status(file_path, sheet_name='입고 스케줄_상세'):
    """입고 스케줄_상세 시트에서 품목코드별 PLC 상태(운영/단종) 추출.
    제품명 셀에 취소선(strikethrough) 서식이 적용된 행은 단종으로 간주 (파일 작성자가 수기로 표시하는 규칙).
    """
    if not os.path.exists(file_path):
        return {}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"   > ⚠️ PLC 상태 파싱 실패: {e}")
        return {}
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    header_row, code_col, name_col = None, None, None
    for r in range(1, min(7, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() == '품목코드':
                header_row, code_col = r, c
            if v is not None and str(v).strip() == '제품명':
                name_col = c
        if header_row is not None:
            break
    if header_row is None or code_col is None or name_col is None:
        print("   > ⚠️ PLC 상태: 품목코드/제품명 헤더를 찾지 못함")
        return {}

    status_map = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(row=r, column=code_col).value
        if code is None:
            continue
        code = str(code).strip().upper()
        if not code or code == '품목코드':
            continue
        font = ws.cell(row=r, column=name_col).font
        status_map[code] = '단종' if (font and font.strike) else '운영'
    return status_map


def parse_forecast_sales(file_path, base_year=2026):
    """포캐스팅 파일의 'N월_판매예정' 탭들에서 SKU별 US/UK/EU 월별 판매예정 수량 파싱.

    시트 구조: 1행=부서, 2행=채널명, 3행=헤더('품목코드'/'수량'/'매출원가'), 4행부터 데이터.
    채널명이 '미국*'이면 US, '영국*'이면 UK, 'EU*'이면 EU로 합산 (MX/일본/글로벌 등 나머지 채널 제외).
    시트명 'N월_판매예정'은 base_year, 'NN년 N월_판매예정'은 20NN년으로 해석.

    반환: ({품목코드(upper): {"YYYY-MM": {"us":n, "uk":n, "eu":n}}}, [월 키 정렬 리스트])
    """
    result = {}
    months_found = set()
    if not os.path.exists(file_path):
        print(f"   > ⚠️ 포캐스팅 파일 없음: {file_path}")
        return result, []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    pat = re.compile(r'^(?:(\d{2})년\s*)?(\d{1,2})월_판매예정$')
    for sheet in wb.sheetnames:
        m = pat.match(sheet.strip())
        if not m:
            continue
        year = 2000 + int(m.group(1)) if m.group(1) else base_year
        month_key = f"{year}-{int(m.group(2)):02d}"
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)          # 1행: 부서
            r2 = next(rows)     # 2행: 채널명
            r3 = next(rows)     # 3행: 수량/매출원가 헤더
        except StopIteration:
            continue
        # 품목코드 컬럼 위치 (통상 A열이지만 방어적으로 탐색)
        code_idx = 0
        for i, v in enumerate(r3):
            if str(v).strip() == '품목코드':
                code_idx = i
                break
        region_cols = {'us': [], 'uk': [], 'eu': []}
        for i, name in enumerate(r2):
            if not name or i >= len(r3) or str(r3[i]).strip() != '수량':
                continue
            nm = str(name).strip()
            if nm.startswith('미국'):
                region_cols['us'].append(i)
            elif nm.startswith('영국'):
                region_cols['uk'].append(i)
            elif nm.startswith('EU'):
                region_cols['eu'].append(i)
        if not any(region_cols.values()):
            print(f"   > ⚠️ '{sheet}' 탭에서 미국/영국/EU 수량 컬럼을 찾지 못해 건너뜀")
            continue
        months_found.add(month_key)
        for row in rows:
            if not row or code_idx >= len(row) or row[code_idx] is None:
                continue
            code = str(row[code_idx]).strip().upper()
            if not code:
                continue
            entry = result.setdefault(code, {}).setdefault(month_key, {'us': 0, 'uk': 0, 'eu': 0})
            for reg, cols in region_cols.items():
                s = 0.0
                for c in cols:
                    v = row[c] if c < len(row) else None
                    try:
                        s += float(v or 0)
                    except (TypeError, ValueError):
                        pass
                entry[reg] += int(round(s))
    return result, sorted(months_found)


def build_pallet_map(path='마스터 계정.xlsx', sheet_name='패킹리스트'):
    """마스터 계정.xlsx의 '패킹리스트' 탭에서 품목코드별 팔렛트당 개입수 추출.
    헤더는 2행 고정, A열=품목코드, N열=아웃박스 입수량, V열=1PLT 카톤박스 총 적재수량.
    팔렛트당 개입수 = N × V.
    """
    if not os.path.exists(path):
        return {}
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=1)
    except Exception as e:
        print(f"   > ⚠️ 팔렛트 환산표 파싱 실패: {e}")
        return {}

    code_col = df.columns[0]
    box_col = next((c for c in df.columns if '아웃박스' in str(c) and '입수' in str(c)), None)
    plt_col = next((c for c in df.columns if '1PLT' in str(c) and '적재' in str(c)), None)
    if box_col is None or plt_col is None:
        print(f"   > ⚠️ 팔렛트 환산표 컬럼 못 찾음 (box={box_col}, plt={plt_col})")
        return {}

    pallet_map = {}
    for _, row in df.iterrows():
        code = row[code_col]
        if pd.isna(code):
            continue
        code = str(code).strip().upper()
        if not code:
            continue
        box_qty = parse_int(row[box_col])
        plt_qty = parse_int(row[plt_col])
        units = box_qty * plt_qty
        if units > 0:
            pallet_map[code] = units
    return pallet_map


def build_b2b_map(prefix='B2B할당_'):
    """B2B 사전할당 소형 엑셀(품목코드 + B2B물량 총량, 파일명에 날짜 포함) 자동 탐색·파싱.
    파일이 아직 없는 주에는 빈 딕셔너리를 반환해 정상 동작을 유지한다.
    """
    path = find_latest_inventory_file(prefix)
    if path is None:
        return {}, None
    df = load_robust_excel(path, '품목코드')
    if df is None or '품목코드' not in df.columns:
        print(f"   > ⚠️ B2B 사전할당 파일({path})에서 '품목코드' 컬럼을 찾지 못함")
        return {}, path

    qty_col = next((c for c in df.columns if 'B2B' in str(c) and ('물량' in str(c) or '수량' in str(c))), None)
    if qty_col is None:
        print(f"   > ⚠️ B2B 사전할당 파일({path})에서 물량 컬럼을 찾지 못함")
        return {}, path

    b2b_map = {}
    for _, row in df.dropna(subset=['품목코드']).iterrows():
        code = str(row['품목코드']).strip().upper()
        if not code:
            continue
        qty = parse_int(row[qty_col])
        if qty > 0:
            b2b_map[code] = b2b_map.get(code, 0) + qty
    return b2b_map, path


def build_master_lookup(master_df):
    name_dict = {}
    if master_df is None or '품목코드' not in master_df.columns or '제품명' not in master_df.columns:
        return name_dict
    for _, row in master_df.dropna(subset=['품목코드']).iterrows():
        raw_code = str(row['품목코드']).strip()
        name = str(row['제품명']).strip() if pd.notna(row.get('제품명')) else ""
        if not name or not raw_code:
            continue
        if ',' in raw_code:
            continue
        if name_dict.get(raw_code, "") == "":
            name_dict[raw_code] = name
    return name_dict


def lookup_product_name(code, name_dict):
    code_str = str(code).strip()
    if ',' in code_str:
        return ""
    return name_dict.get(code_str, "")


def apply_sales_plc_override(df):
    """PLC 자동판정(취소선 기반) 보정: 지역 합계 판매량이 100개 이상이면 소스 파일 표시와 무관하게 '운영'으로 간주.
    실제로 잘 팔리고 있는데도 원본 파일의 취소선 표시(오기입 등) 때문에 '단종'으로 잘못 뜨는 사례를 보정하기 위함.
    """
    total_sales = df['us_sales'] + df['uk_sales'] + df['eu_sales']
    df.loc[(total_sales >= 100) & (df['plc_status'] != '운영'), 'plc_status'] = '운영'


def build_excluded_codes(master_df):
    """마스터 계정.xlsx의 '제외' 컬럼에 표시된 품목코드 집합 반환 (단종/미운영 등 리포트 제외 대상)"""
    excluded = set()
    if master_df is None or '제외' not in master_df.columns or '품목코드' not in master_df.columns:
        return excluded
    for _, row in master_df.dropna(subset=['품목코드']).iterrows():
        flag = str(row.get('제외', '')).strip().upper()
        if flag in ('Y', 'TRUE', '1', '제외', 'O'):
            raw_code = str(row['품목코드']).strip()
            if ',' in raw_code:
                for c in raw_code.split(','):
                    if c.strip():
                        excluded.add(c.strip().upper())
            else:
                excluded.add(raw_code.upper())
    return excluded


# 스냅샷 델타 비교에 쓰는 필드 (전체 필드를 임베드하면 파일이 비대해지므로 화면 델타에 쓰는 것만)
SNAPSHOT_DELTA_FIELDS = ['us_total', 'us_days', 'us_sales',
                         'uk_total', 'uk_days', 'uk_sales',
                         'eu_total', 'eu_days', 'eu_sales']


def update_snapshots(json_tab1, json_tab2, snap_dir="snapshots", keep_days=30):
    """계산 완료된 결과(rawData/groupedData)를 날짜별로 동결 저장하고,
    최근 keep_days일치에서 델타 비교용 필드만 추려 history 객체로 반환한다.

    - full_YYYY-MM-DD.json: 그날의 전체 계산 결과 동결본. 같은 날 재실행 시 그날 최종 상태로 갱신.
      (다른 날짜는 절대 덮어쓰지 않음 — 과거 시점 박제가 목적.)
    - 판매가 나중에 소급 재집계돼도 과거 동결본은 안 바뀜 → '그날 대시보드가 뭘 위험이라 판정했나'가 보존됨.
    """
    os.makedirs(snap_dir, exist_ok=True)
    today = datetime.now().date()
    today_key = today.isoformat()

    # 1) 오늘 동결본 저장 (전체 필드)
    full_path = os.path.join(snap_dir, f"full_{today_key}.json")
    with open(full_path, 'w', encoding='utf-8') as f:
        json.dump({"date": today_key, "generated_at": datetime.now().isoformat(),
                   "tab1": json_tab1, "tab2": json_tab2}, f, ensure_ascii=False)

    # 2) 보관 기간 지난 동결본 삭제
    cutoff = today - timedelta(days=keep_days)
    for p in glob.glob(os.path.join(snap_dir, "full_*.json")):
        m = re.search(r'full_(\d{4}-\d{2}-\d{2})\.json$', p.replace('\\', '/'))
        if m:
            try:
                if datetime.strptime(m.group(1), "%Y-%m-%d").date() < cutoff:
                    os.remove(p)
            except ValueError:
                pass

    # 3) 남은 동결본에서 델타 필드만 추려 history 구성 (dashboard.html에 임베드용)
    hist = {"dates": [], "tab1": {}, "tab2": {}}
    for p in sorted(glob.glob(os.path.join(snap_dir, "full_*.json"))):
        m = re.search(r'full_(\d{4}-\d{2}-\d{2})\.json$', p.replace('\\', '/'))
        if not m:
            continue
        dkey = m.group(1)
        try:
            with open(p, encoding='utf-8') as f:
                snap = json.load(f)
        except Exception:
            continue
        hist["dates"].append(dkey)
        for tab_key in ("tab1", "tab2"):
            hist[tab_key][dkey] = {
                str(r.get("code")): {fld: r.get(fld) for fld in SNAPSHOT_DELTA_FIELDS if fld in r}
                for r in snap.get(tab_key, [])
            }
    print(f"   > 스냅샷: {today_key} 동결 저장, 비교 가능 이력 {len(hist['dates'])}일치 ({hist['dates'][0] if hist['dates'] else '-'} ~ {hist['dates'][-1] if hist['dates'] else '-'})")
    return hist


def build_heavy_codes(master_df):
    """마스터 계정.xlsx의 '중량물' 컬럼에 표시된 품목코드 집합 반환.
    중량물은 항공 운임이 매우 비싸 안전재고 일수 +30일, 항공 경고 임계선 +30일을 적용한다 (2026-07-15).
    컬럼이 없으면 빈 집합 — 기능이 조용히 꺼진 상태로 기존과 동일하게 동작."""
    heavy = set()
    if master_df is None or '중량물' not in master_df.columns or '품목코드' not in master_df.columns:
        return heavy
    for _, row in master_df.dropna(subset=['품목코드']).iterrows():
        flag = str(row.get('중량물', '')).strip().upper()
        if flag in ('Y', 'TRUE', '1', '중량', '중량물', 'O'):
            raw_code = str(row['품목코드']).strip()
            for c in raw_code.split(','):
                if c.strip():
                    heavy.add(c.strip().upper())
    return heavy


HEAVY_SAFETY_EXTRA_DAYS = 30  # 중량물 안전재고 가산 일수 (template.html의 항공 임계선 +30과 세트)


def build_json_rows(df):
    def calc_days(total, sales):
        if sales <= 0:
            return "여유"
        try:
            return str(int(total / sales))
        except:
            return "0"

    for r in ['us', 'uk', 'eu']:
        df[f'{r}_total'] = df[f'{r}_stock'] + df[f'{r}_air'] + df[f'{r}_sea']

    if 'is_heavy' not in df.columns:
        df['is_heavy'] = False
    df['is_heavy'] = df['is_heavy'].fillna(False).astype(bool)
    heavy_extra = df['is_heavy'].astype(int) * HEAVY_SAFETY_EXTRA_DAYS
    df['us_safety'] = df['us_sales'] * (90 + heavy_extra)
    df['uk_safety'] = df['uk_sales'] * (180 + heavy_extra)
    df['eu_safety'] = df['eu_sales'] * (180 + heavy_extra)

    df['us_days'] = df.apply(lambda row: calc_days(row['us_total'], row['us_sales']), axis=1)
    df['uk_days'] = df.apply(lambda row: calc_days(row['uk_total'], row['uk_sales']), axis=1)
    df['eu_days'] = df.apply(lambda row: calc_days(row['eu_total'], row['eu_sales']), axis=1)

    json_data = []
    for _, row in df.iterrows():
        json_data.append({
            "code": str(row['품목코드']),
            "name": str(row['제품명']),
            "상품라인": str(row.get('상품라인', '기타(미등록)')),
            "통합 제품명": str(row.get('통합 제품명', '미등록')),
            "us_stock": int(row['us_stock']),
            "us_air": int(row['us_air']),
            "us_sea": int(row['us_sea']),
            "us_total": int(row['us_total']),
            "us_sales": int(row['us_sales']),
            "us_safety": int(row['us_safety']),
            "us_days": str(row['us_days']),
            "uk_stock": int(row['uk_stock']),
            "uk_air": int(row['uk_air']),
            "uk_sea": int(row['uk_sea']),
            "uk_total": int(row['uk_total']),
            "uk_sales": int(row['uk_sales']),
            "uk_safety": int(row['uk_safety']),
            "uk_days": str(row['uk_days']),
            "eu_stock": int(row['eu_stock']),
            "eu_air": int(row['eu_air']),
            "eu_sea": int(row['eu_sea']),
            "eu_total": int(row['eu_total']),
            "eu_sales": int(row['eu_sales']),
            "eu_safety": int(row['eu_safety']),
            "eu_days": str(row['eu_days']),
            "plc_status": str(row.get('plc_status', '정보없음')),
            "us_value": int(row.get('us_value', 0)),
            "uk_value": int(row.get('uk_value', 0)),
            "eu_value": int(row.get('eu_value', 0)),
            "stock_value": int(row.get('stock_value', 0)),
            "pallet_units": int(row.get('pallet_units', 0)),
            "b2b_qty": int(row.get('b2b_qty', 0)),
            "kr_stock": int(row.get('kr_stock', 0)),
            "is_heavy": bool(row.get('is_heavy', False)),
            "us_spike": bool(row.get('us_spike', False)),
            "uk_spike": bool(row.get('uk_spike', False)),
            "eu_spike": bool(row.get('eu_spike', False)),
            "us_sales_window": int(row.get('us_sales_window', 28) or 28),
            "uk_sales_window": int(row.get('uk_sales_window', 28) or 28),
            "eu_sales_window": int(row.get('eu_sales_window', 28) or 28),
            "us_last7": [int(x) for x in (row.get('us_last7') if isinstance(row.get('us_last7'), list) else [0]*7)],
            "uk_last7": [int(x) for x in (row.get('uk_last7') if isinstance(row.get('uk_last7'), list) else [0]*7)],
            "eu_last7": [int(x) for x in (row.get('eu_last7') if isinstance(row.get('eu_last7'), list) else [0]*7)],
            "fcst": row.get('fcst') if isinstance(row.get('fcst'), dict) else {}
        })
    return json_data


def run_dashboard_generator():
    # 파일 경로 — 날짜 자동 탐색
    file_gen = find_latest_inventory_file("일반_전체재고_")
    file_fbt = find_latest_inventory_file("FBT_전체재고_")
    file_fba = find_latest_inventory_file("FBA_전체재고_")

    excel_us_sales = find_latest_sales_file(["미국"])
    excel_uk_sales = find_latest_sales_file(["영국", "UK"])
    excel_eu_sales = find_latest_sales_file(["EU"])
    excel_transit  = "2025_이동중_재고_현황.xlsx"
    excel_master   = "마스터 계정.xlsx"
    excel_inbound  = "2026_SCM_영업채널별_판매_포캐스팅_공유용.xlsx"

    template_path  = "template.html"
    output_path    = "dashboard.html"

    for name, path in [('일반재고', file_gen), ('FBT재고', file_fbt), ('FBA재고', file_fba),
                       ('US판매', excel_us_sales), ('UK판매', excel_uk_sales), ('EU판매', excel_eu_sales)]:
        if path is None:
            print(f"❌ {name} 파일을 찾을 수 없습니다.")
            return
        print(f"   > {name}: {path} 자동 선택")

    print("🔄 [1/7] 마스터 계정 기반 품목코드->제품명 사전 구축 중...")
    master_df = load_robust_excel(excel_master, '품목코드')
    global_name_dict = build_master_lookup(master_df)
    print(f"   > {len(global_name_dict)}개 단일코드-제품명 매핑 완료")
    excluded_codes = build_excluded_codes(master_df)
    print(f"   > 마스터 계정 '제외' 표시: {len(excluded_codes)}개 품목코드 리포트 제외 대상")
    heavy_codes = build_heavy_codes(master_df)
    if heavy_codes:
        print(f"   > 마스터 계정 '중량물' 표시: {len(heavy_codes)}개 품목코드 — 안전재고/항공 임계선 +{HEAVY_SAFETY_EXTRA_DAYS}일 적용")
    else:
        print("   > 마스터 계정에 '중량물' 컬럼 없음(또는 표시 0건) — 중량물 차등 미적용")

    print("🔄 [2/7] 통합 재고 파일(일반/FBT/FBA) 읽기 중...")
    stock_df = load_integrated_inventory(file_gen, file_fbt, file_fba)

    print("🔄 [3/7] 국가별 일판매 데이터 추출 중 (28일 롤링 평균)...")
    us_sales_df, us_daily = read_sales_rolling(excel_us_sales, "us")
    uk_sales_df, uk_daily = read_sales_rolling(excel_uk_sales, "uk")
    eu_sales_df, eu_daily = read_sales_rolling(excel_eu_sales, "eu")

    print("🔄 [4/7] 이동중(In-Transit) 합산 데이터 결합 중...")
    transit_df = read_transit_data(excel_transit)

    print("🔄 [5/7] ETA(현지 입고 일정) 파일 파싱 중 (US/UK 항공·해상)...")
    eta_schedule = parse_transit_schedule(excel_transit, year=2026)

    print("🔄 [5.5/7] 국내 메인창고 입고 스케줄 파싱 중...")
    domestic_inbound = parse_domestic_inbound(excel_inbound)

    print("🔄 [5.6/7] 원가/PLC 상태 파싱 중...")
    product_cost = parse_product_cost(excel_inbound)
    product_status = parse_product_status(excel_inbound)
    print(f"   > 원가 매칭 가능 SKU: {len(product_cost)}개, PLC 상태 정보: {len(product_status)}개")

    print("🔄 [5.65/7] 월별 판매예정(포캐스팅) 파싱 중...")
    fcst_map, fcst_months = parse_forecast_sales(excel_inbound)
    print(f"   > 판매예정: {len(fcst_months)}개월 ({', '.join(fcst_months)}), {len(fcst_map)}개 코드")

    # 포캐스팅에는 수량이 있는데 마스터 계정에 없는 코드 경고.
    # 제품 리뉴얼로 코드가 바뀌면(GLMX124 → GLMX124_V02) 포캐스팅은 신코드로 잡히는데
    # 마스터에 신코드가 없으면 그 물량이 대시보드/워터폴에서 조용히 0으로 빠진다 (2026-07-22 발견).
    fcst_unmapped = []
    if master_df is not None and '품목코드' in master_df.columns:
        master_code_set = set()
        for c in master_df['품목코드'].dropna().astype(str):
            for p in c.split(','):
                if p.strip():
                    master_code_set.add(p.strip().upper())
        for code, months in fcst_map.items():
            qty = sum(sum(v.values()) for v in months.values())
            if qty > 0 and code not in master_code_set:
                fcst_unmapped.append((code, qty))
        fcst_unmapped.sort(key=lambda x: -x[1])
    if fcst_unmapped:
        total_missing = sum(q for _, q in fcst_unmapped)
        print(f"   > ⚠️ 포캐스팅 물량이 있으나 마스터 계정에 없는 코드: {len(fcst_unmapped)}개 "
              f"(합계 {total_missing:,}개) — 이 물량은 워터폴/포캐스팅 비교에서 누락됨")
        print(f"     예시: {', '.join(c for c, _ in fcst_unmapped[:10])}"
              f"{' 외 다수' if len(fcst_unmapped) > 10 else ''}")
        print(f"     ↳ 해결: 마스터 계정.xlsx에 해당 코드를 추가하고 '통합 제품명'을 기존 제품과 동일하게 지정")

    print("🔄 [5.7/7] 부킹 관리용 팔렛트 환산표/B2B 사전할당/한국 가용재고 파싱 중...")
    pallet_map = build_pallet_map(excel_master)
    print(f"   > 팔렛트 환산표(패킹리스트): {len(pallet_map)}개 코드 매칭")
    kr_stock_map = parse_kr_available(file_gen)
    print(f"   > 한국(콜로세움) 가용재고: {len(kr_stock_map)}개 코드")
    b2b_map, b2b_path = build_b2b_map("B2B할당_")
    if b2b_path:
        print(f"   > B2B 사전할당: {b2b_path} 자동 선택, {len(b2b_map)}개 코드")
    else:
        print("   > B2B 사전할당: 파일 없음 (전량 0으로 처리)")

    # 데이터 병합
    merged = stock_df.copy()
    if merged.empty:
        merged = pd.DataFrame(columns=['품목코드', 'us_stock', 'uk_stock', 'eu_stock'])

    merged = pd.merge(merged, us_sales_df, on='품목코드', how='outer')
    merged = pd.merge(merged, uk_sales_df, on='품목코드', how='outer')
    merged = pd.merge(merged, eu_sales_df, on='품목코드', how='outer')
    merged = pd.merge(merged, transit_df,  on='품목코드', how='outer')

    fill_numeric = ['us_stock', 'uk_stock', 'eu_stock',
                    'us_sales', 'uk_sales', 'eu_sales',
                    'us_air', 'us_sea', 'uk_air', 'uk_sea', 'eu_air', 'eu_sea']
    for c in fill_numeric:
        if c not in merged.columns:
            merged[c] = 0
    merged[fill_numeric] = merged[fill_numeric].fillna(0)

    # 28일 롤링에서 나온 급증 플래그/최근7일 컬럼 결측 보정 (재고만 있고 판매 데이터 없는 SKU)
    for r in ['us', 'uk', 'eu']:
        spike_col, last7_col = f'{r}_spike', f'{r}_last7'
        if spike_col not in merged.columns:
            merged[spike_col] = False
        merged[spike_col] = merged[spike_col].apply(lambda v: bool(v) if pd.notna(v) else False)
        if last7_col not in merged.columns:
            merged[last7_col] = None
        merged[last7_col] = merged[last7_col].apply(lambda v: v if isinstance(v, list) else [0]*7)
        window_col = f'{r}_sales_window'
        if window_col not in merged.columns:
            merged[window_col] = 28
        merged[window_col] = merged[window_col].fillna(28).astype(int)

    merged['품목코드'] = merged['품목코드'].astype(str).str.strip()
    merged['제품명'] = merged['품목코드'].apply(lambda c: lookup_product_name(c, global_name_dict))

    if excluded_codes:
        before_count = len(merged)
        merged = merged[~merged['품목코드'].str.upper().isin(excluded_codes)]
        print(f"   > 제외 대상 필터링: {before_count - len(merged)}개 행 제외됨")

    # 원가/PLC 상태 부착 + 재고 금액 계산 (그룹핑 전, 개별 코드 단위로 계산해야 통합 제품명 합산 시 정확함)
    merged['unit_cost'] = merged['품목코드'].str.upper().map(product_cost).fillna(0).astype(int)
    merged['plc_status'] = merged['품목코드'].str.upper().map(product_status).fillna('정보없음')
    merged['pallet_units'] = merged['품목코드'].str.upper().map(pallet_map).fillna(0).astype(int)
    merged['b2b_qty'] = merged['품목코드'].str.upper().map(b2b_map).fillna(0).astype(int)
    merged['kr_stock'] = merged['품목코드'].str.upper().map(kr_stock_map).fillna(0).astype(int)
    merged['is_heavy'] = merged['품목코드'].str.upper().isin(heavy_codes)
    # 월별 판매예정(포캐스팅) 부착 — 코드 단위. tab2에서는 그룹 내 코드끼리 월별/권역별 합산됨.
    merged['fcst'] = merged['품목코드'].str.upper().map(lambda c: fcst_map.get(c, {}))
    for r in ['us', 'uk', 'eu']:
        merged[f'{r}_total'] = merged[f'{r}_stock'] + merged[f'{r}_air'] + merged[f'{r}_sea']
        merged[f'{r}_value'] = merged[f'{r}_total'] * merged['unit_cost']
    merged['stock_value'] = merged['us_value'] + merged['uk_value'] + merged['eu_value']

    print("🔄 [6/7] 마스터 계정 연동 및 그룹 정보 부착 중...")
    if master_df is not None:
        cols_to_use = ['품목코드']
        if '상품라인'   in master_df.columns: cols_to_use.append('상품라인')
        if '통합 제품명' in master_df.columns: cols_to_use.append('통합 제품명')

        master_mapping = master_df[cols_to_use].drop_duplicates(subset=['품목코드'])
        master_mapping['품목코드'] = master_mapping['품목코드'].astype(str).str.strip()

        merged['품목코드_str'] = merged['품목코드'].astype(str).str.strip()
        merged = pd.merge(merged, master_mapping,
                          left_on='품목코드_str', right_on='품목코드',
                          how='left', suffixes=('', '_master'))

        merged['상품라인']   = merged['상품라인'].fillna('기타(미등록)')   if '상품라인'   in merged.columns else '기타(미등록)'
        merged['통합 제품명'] = merged['통합 제품명'].fillna('미등록') if '통합 제품명' in merged.columns else '미등록'
    else:
        merged['상품라인']   = '기타(미등록)'
        merged['통합 제품명'] = '미등록'

    merged = merged[merged['제품명'].str.strip() != ""]

    print("🔄 [7/7] JSON 생성 및 HTML 발행 중...")

    tab1 = merged.copy()
    tab1['grand_total'] = tab1['us_stock'] + tab1['uk_stock'] + tab1['eu_stock'] + \
                          tab1['us_air'] + tab1['us_sea'] + \
                          tab1['uk_air'] + tab1['uk_sea'] + \
                          tab1['eu_air'] + tab1['eu_sea']
    # 재고+이동중이 0이어도 판매가 발생 중이거나 국내 재고가 있으면 표시 —
    # 결품 상태로 팔리는 SKU가 대시보드에서 아예 사라져 항공 긴급을 놓치는 문제 방지 (2026-07-15)
    tab1_sales = tab1['us_sales'] + tab1['uk_sales'] + tab1['eu_sales']
    stockout_selling = ((tab1['grand_total'] <= 0) & (tab1_sales > 0)).sum()
    tab1 = tab1[(tab1['grand_total'] > 0) | (tab1_sales > 0) | (tab1['kr_stock'] > 0)]
    if stockout_selling:
        print(f"   > 🚨 결품 판매중 SKU {stockout_selling}개 — 재고 0이지만 판매 발생, 대시보드에 표시됨")

    # PLC 자동판정(단종) 보정: 판매가 실제로 발생 중이면(합계 100개 이상) 소스 파일 표시와 무관하게 운영으로 간주
    apply_sales_plc_override(tab1)

    unmatched_codes = tab1.loc[tab1['unit_cost'] <= 0, '품목코드'].tolist()
    print(f"   > ⚠️ 원가 미매칭 SKU: {len(unmatched_codes)}개 (재고 금액 계산에서 제외됨)")
    if unmatched_codes:
        preview = ', '.join(unmatched_codes[:15])
        print(f"     예시: {preview}{' 외 다수' if len(unmatched_codes) > 15 else ''}")

    unmatched_pallet_codes = tab1.loc[tab1['pallet_units'] <= 0, '품목코드'].tolist()
    print(f"   > ⚠️ 팔렛트 환산표 미매칭 SKU: {len(unmatched_pallet_codes)}개 (부킹 관리에서 PLT 계산 제외됨)")
    if unmatched_pallet_codes:
        preview = ', '.join(unmatched_pallet_codes[:15])
        print(f"     예시: {preview}{' 외 다수' if len(unmatched_pallet_codes) > 15 else ''}")

    sort_cols = [c for c in ['상품라인', '통합 제품명'] if c in tab1.columns]
    if sort_cols:
        tab1 = tab1.sort_values(by=sort_cols).reset_index(drop=True)

    json_tab1 = build_json_rows(tab1.copy())

    sum_cols  = ['us_stock', 'uk_stock', 'eu_stock',
                 'us_sales', 'uk_sales', 'eu_sales',
                 'us_air', 'us_sea', 'uk_air', 'uk_sea', 'eu_air', 'eu_sea',
                 'us_value', 'uk_value', 'eu_value', 'stock_value', 'kr_stock']
    # 그룹 키는 '통합 제품명' 단독으로 사용한다.
    # 과거에는 ['상품라인', '통합 제품명']을 함께 키로 썼는데, 마스터 계정.xlsx에 같은 통합
    # 제품명인데 상품라인이 다르게(오타/공백) 입력된 그룹이 있으면 그 통합 제품명이 여러 개의
    # 파편 그룹으로 쪼개져 재고 합계·PLC 상태가 일부 SKU만 반영되는 문제가 있었음(예: GLMX179/GLMX191).
    group_key = ['통합 제품명']

    # PLC: 그룹 내 아무 코드나 '운영'이면 그룹 전체를 '운영'으로 본다 (마스터 계정 순서 기반 판정은
    # 마스터 데이터 정합성에 너무 민감해서 보류 — 2026-07-01 재논의 후 단순 규칙으로 되돌림)
    def agg_plc_status(values):
        vals = set(values)
        if '운영' in vals:
            return '운영'
        if '단종' in vals:
            return '단종'
        return '정보없음'

    def agg_last7(series):
        # 그룹 내 코드들의 최근 7일 리스트를 요소별 합산
        lists = [v for v in series if isinstance(v, list)]
        if not lists:
            return [0] * 7
        return [sum(vals) for vals in zip(*lists)]

    def agg_fcst(series):
        # 그룹 내 코드들의 월별 판매예정을 월/권역별로 합산
        out = {}
        for d in series:
            if not isinstance(d, dict):
                continue
            for mk, regs in d.items():
                o = out.setdefault(mk, {'us': 0, 'uk': 0, 'eu': 0})
                for r in ('us', 'uk', 'eu'):
                    o[r] += regs.get(r, 0)
        return out

    merged_rep = merged.groupby(group_key, sort=False).agg(
        대표제품명=('제품명', 'first'),
        대표상품라인=('상품라인', 'first'),
        plc_status=('plc_status', agg_plc_status),
        us_spike=('us_spike', 'any'),
        uk_spike=('uk_spike', 'any'),
        eu_spike=('eu_spike', 'any'),
        is_heavy=('is_heavy', 'any'),
        us_last7=('us_last7', agg_last7),
        uk_last7=('uk_last7', agg_last7),
        eu_last7=('eu_last7', agg_last7),
        us_sales_window=('us_sales_window', 'min'),
        uk_sales_window=('uk_sales_window', 'min'),
        eu_sales_window=('eu_sales_window', 'min'),
        fcst=('fcst', agg_fcst),
    ).reset_index()
    merged_sum = merged.groupby(group_key, sort=False)[sum_cols].sum().reset_index()

    tab2 = pd.merge(merged_rep, merged_sum, on=group_key)

    # 그룹 포캐스팅도 마스터 매핑 기준으로 재계산 — merged에 없는 신코드(재고 미도착) 물량 포함
    group_fcst = build_group_fcst(master_df, fcst_map)
    if group_fcst:
        before_zero = sum(1 for v in tab2['fcst'] if not v)
        tab2['fcst'] = tab2['통합 제품명'].map(lambda g: group_fcst.get(str(g).strip(), {}))
        after_zero = sum(1 for v in tab2['fcst'] if not v)
        print(f"   > 그룹 포캐스팅: 마스터 매핑 기준 재집계 ({len(group_fcst)}개 그룹, 빈 그룹 {before_zero}→{after_zero})")

    # 그룹 일판매는 코드별 값 합산이 아니라 '일자별 합계'에서 재계산한다.
    # (코드별 채택 윈도우가 다르면 서로 다른 기간의 수요가 더해져 과대 계상됨 — build_group_sales 주석 참고)
    group_sales = build_group_sales(merged, {'us': us_daily, 'uk': uk_daily, 'eu': eu_daily})
    for r, gdf in group_sales.items():
        cols = [f'{r}_sales', f'{r}_sales_window', f'{r}_last7', f'{r}_spike']
        tab2 = tab2.drop(columns=[c for c in cols if c in tab2.columns])
        tab2 = pd.merge(tab2, gdf, on='통합 제품명', how='left')
        tab2[f'{r}_sales'] = tab2[f'{r}_sales'].fillna(0).astype(int)
        tab2[f'{r}_sales_window'] = tab2[f'{r}_sales_window'].fillna(28).astype(int)
        tab2[f'{r}_spike'] = tab2[f'{r}_spike'].apply(lambda v: bool(v) if pd.notna(v) else False)
        tab2[f'{r}_last7'] = tab2[f'{r}_last7'].apply(lambda v: v if isinstance(v, list) else [0]*7)

    tab2['품목코드'] = tab2['통합 제품명']
    tab2 = tab2.rename(columns={'대표제품명': '제품명', '대표상품라인': '상품라인'})

    tab2['grand_total'] = tab2['us_stock'] + tab2['uk_stock'] + tab2['eu_stock'] + \
                          tab2['us_air'] + tab2['us_sea'] + \
                          tab2['uk_air'] + tab2['uk_sea'] + \
                          tab2['eu_air'] + tab2['eu_sea']
    tab2_sales = tab2['us_sales'] + tab2['uk_sales'] + tab2['eu_sales']
    tab2 = tab2[(tab2['grand_total'] > 0) | (tab2_sales > 0) | (tab2['kr_stock'] > 0)]

    apply_sales_plc_override(tab2)

    if sort_cols:
        tab2 = tab2.sort_values(by=sort_cols).reset_index(drop=True)

    json_tab2 = build_json_rows(tab2.copy())

    # 워터폴 데이터 (현지 ETA)
    waterfall_map = {}
    if '통합 제품명' in merged.columns:
        for unified_name, group in merged.groupby('통합 제품명'):
            if not unified_name or unified_name == '미등록':
                continue
            schedules = []
            for sku in group['품목코드'].astype(str).str.strip().str.upper().unique():
                if sku in eta_schedule:
                    schedules.extend(eta_schedule[sku])
            if schedules:
                waterfall_map[unified_name] = schedules
    print(f"   > 워터폴 데이터: {len(waterfall_map)}개 통합 제품명에 입고 일정 매핑")

    # 국내 입고 데이터 (통합 제품명별 합산)
    inbound_map = {}
    if '통합 제품명' in merged.columns:
        for unified_name, group in merged.groupby('통합 제품명'):
            if not unified_name or unified_name == '미등록':
                continue
            agg_total = 0
            agg = {}
            for sku in group['품목코드'].astype(str).str.strip().str.upper().unique():
                if sku not in domestic_inbound:
                    continue
                entry = domestic_inbound[sku]
                agg_total += entry['total']
                for it in entry['items']:
                    key = (it['month'], it['week'])
                    agg[key] = agg.get(key, 0) + it['qty']
            if agg_total > 0 or agg:
                items = [{'month': k[0], 'week': k[1], 'qty': v} for k, v in agg.items()]
                inbound_map[unified_name] = {'total': agg_total, 'items': items}
    print(f"   > 국내 입고 예정: {len(inbound_map)}개 통합 제품명에 매핑 완료")

    # 일별 판매 추이 (통합 제품명별, 최근 28일) — 모달의 '일별 판매' 탭용
    today = datetime.now().date()
    win_dates = [today - timedelta(days=28) + timedelta(days=i) for i in range(28)]
    region_daily = {'us': us_daily, 'uk': uk_daily, 'eu': eu_daily}
    daily_sales_map = {}
    if '통합 제품명' in merged.columns:
        for unified_name, group in merged.groupby('통합 제품명'):
            if not unified_name or unified_name == '미등록':
                continue
            skus = group['품목코드'].astype(str).str.strip().str.upper().unique()
            entry = {}
            has_any = False
            for r, dmap in region_daily.items():
                series = [0] * 28
                for sku in skus:
                    d = dmap.get(sku)
                    if not d:
                        continue
                    for i, dt in enumerate(win_dates):
                        q = d.get(dt, 0)
                        if q:
                            series[i] += q
                if any(series):
                    has_any = True
                entry[r] = series
            if has_any:
                entry['dates'] = [d.isoformat() for d in win_dates]
                daily_sales_map[unified_name] = entry
    print(f"   > 일별 판매 추이: {len(daily_sales_map)}개 통합 제품명에 매핑 완료")

    cost_meta = {
        "matched_count": int((tab1['unit_cost'] > 0).sum()),
        "unmatched_count": len(unmatched_codes),
        "unmatched_codes": unmatched_codes
    }

    pallet_meta = {
        "matched_count": int((tab1['pallet_units'] > 0).sum()),
        "unmatched_count": len(unmatched_pallet_codes),
        "unmatched_codes": unmatched_pallet_codes,
        "b2b_file": b2b_path,
        "b2b_matched_count": len(b2b_map)
    }

    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
    else:
        print("❌ template.html 파일이 없습니다.")
        return

    html_content = html_content.replace(
        "__DATA_FROM_PYTHON__",
        json.dumps(json_tab1, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__GROUPED_DATA_FROM_PYTHON__",
        json.dumps(json_tab2, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__WATERFALL_DATA_FROM_PYTHON__",
        json.dumps(waterfall_map, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__INBOUND_DATA_FROM_PYTHON__",
        json.dumps(inbound_map, ensure_ascii=False)
    )
    # 부킹 관리용: SKU(개별 코드) 단위 국내 입고 예정 (그룹 단위 inbound_map을 쓰면 그룹 내 코드마다 중복 계산됨)
    html_content = html_content.replace(
        "__INBOUND_SKU_DATA_FROM_PYTHON__",
        json.dumps(domestic_inbound, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__LAST_UPDATED_FROM_PYTHON__",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    )
    html_content = html_content.replace(
        "__COST_META_FROM_PYTHON__",
        json.dumps(cost_meta, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__DAILY_SALES_FROM_PYTHON__",
        json.dumps(daily_sales_map, ensure_ascii=False)
    )
    html_content = html_content.replace(
        "__PALLET_META_FROM_PYTHON__",
        json.dumps(pallet_meta, ensure_ascii=False)
    )
    # 스냅샷 동결 저장 + 최근 30일 델타 비교 이력 임베드
    snapshot_history = update_snapshots(json_tab1, json_tab2)
    html_content = html_content.replace(
        "__SNAPSHOT_HISTORY_FROM_PYTHON__",
        json.dumps(snapshot_history, ensure_ascii=False)
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    # S&OP 리포트(sop_report.py)가 대시보드 HTML을 다시 파싱하지 않도록 데이터만 별도 저장
    dashboard_data_path = "dashboard_data.json"
    with open(dashboard_data_path, 'w', encoding='utf-8') as f:
        json.dump({
            "generated_at": datetime.now().isoformat(),
            "grouped_rows": json_tab2,
            # v9(판매량집계_포캐스팅비교)가 발주/예측에서 단종 SKU를 거르는 데 사용 —
            # 취소선 기반 PLC 판정 + 판매량 보정(apply_sales_plc_override)이 적용된 개별 코드 단위 상태
            "plc_by_code": {str(r['품목코드']).strip().upper(): str(r['plc_status'])
                            for _, r in tab1.iterrows() if str(r.get('plc_status', '')) in ('운영', '단종')}
        }, f, ensure_ascii=False)

    print(f"🎉 성공: '{output_path}' 생성 완료!")


if __name__ == "__main__":
    run_dashboard_generator()